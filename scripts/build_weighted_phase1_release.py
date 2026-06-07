from __future__ import annotations

import csv
import hashlib
import json
import shutil
import zipfile
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from build_visualizations import build_visualizations


ROOT = Path(__file__).resolve().parent.parent
HANDOFF = ROOT / "handoff_package_v2" / "codex_full_handoff_nevada_charter_grad_rates_v2"
RELEASE = ROOT / "release" / "NV_Charter_Grad_Rates_Weighted_Rebuild_v1"
DATE = "2026-06-06"


def short_next_year(year: str) -> str:
    start = int(year[:4])
    end = int(year[5:7])
    return f"{start + 1}-{str(end + 1).zfill(2)}"


def read_csv_dict(path: Path, skip_report_preamble: bool = False) -> list[dict]:
    if not skip_report_preamble:
        with path.open(encoding="utf-8-sig", newline="") as f:
            return list(csv.DictReader(f))
    with path.open(encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    header = rows[3]
    return [dict(zip(header, row)) for row in rows[4:] if row and row[0] != ""]


def write_csv(path: Path, fieldnames: list[str], rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def load_enrollment_counts() -> dict[tuple[str, str], dict]:
    enroll_dir = HANDOFF / "01_SOURCE_DATA" / "Validation_Day_enrollment"
    out: dict[tuple[str, str], dict] = {}
    for path in sorted(enroll_dir.glob("*.xlsx")):
        year = path.name.replace("NDE_Validation_Day_Enrollment_", "").replace("_suppressed", "").replace(".xlsx", "").replace("_", "-")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb["School Grade Subgroups Data"]
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] == "Local Education Agency Code":
                continue
            if row[4] is None or row[6] is None:
                continue
            grade = str(row[6])
            if grade not in {"9", "10", "11", "12"}:
                continue
            school_code = str(row[4])
            school_name = str(row[5])
            total = row[7]
            key = (year, school_code)
            rec = out.setdefault(
                key,
                {
                    "school_name": school_name,
                    "grade_9_count": "",
                    "grade_10_count": "",
                    "grade_11_count": "",
                    "grade_12_count": "",
                    "source_file": path.name,
                    "source_url": "https://doe.nv.gov/offices/office-of-assessment-data-and-accountability-management-adam/accountability/data-requests/enrollment-for-nevada-public-schools",
                },
            )
            rec[f"grade_{grade}_count"] = total
    return out


def load_current_panel() -> list[dict]:
    return read_csv_dict(HANDOFF / "02_CURRENT_REPO_INPUTS" / "nevada_charter_school_year_panel_current.csv")


def load_weighted_summary() -> list[dict]:
    return read_csv_dict(HANDOFF / "03_PRIOR_CALCULATED_OUTPUTS" / "nevada_charter_weighted_summary_all_years.csv")


def load_weighted_audit() -> list[dict]:
    return read_csv_dict(HANDOFF / "03_PRIOR_CALCULATED_OUTPUTS" / "nevada_charter_weighted_line_items_audit.csv")


def load_official_context() -> list[dict]:
    return read_csv_dict(HANDOFF / "03_PRIOR_CALCULATED_OUTPUTS" / "nevada_report_official_context_rates.csv")


def load_source_manifest() -> list[dict]:
    return read_csv_dict(HANDOFF / "01_SOURCE_DATA" / "source_manifest.csv")


def build_universe(panel: list[dict], acgr_export: list[dict], enroll: dict[tuple[str, str], dict]) -> list[dict]:
    rows = []
    overrides = {
        ("2022-23", "Young Women's Leadership Academy of Las Vegas"): ("no", "not_yet_in_graduation_universe", "high"),
        ("2023-24", "Young Women's Leadership Academy of Las Vegas"): ("no", "not_yet_in_graduation_universe", "high"),
        ("2024-25", "Young Women's Leadership Academy of Las Vegas"): ("no", "not_yet_in_graduation_universe", "high"),
        ("2022-23", "Explore Academy"): ("no", "no_grade_12", "high"),
    }
    for r in panel:
        operating_year = r["class_year"]
        reporting_year = short_next_year(operating_year)
        key = (operating_year, r["entity_id"])
        enroll_rec = enroll.get(key, {})
        grade12 = enroll_rec.get("grade_12_count", "")
        has_grade12 = "yes" if (str(grade12).strip() not in {"", "0", "None"} or r["acgr_row_present"] == "yes") else "no"
        include = "yes" if r["include_in_sector_total"] == "yes" and (has_grade12 == "yes" or r["acgr_row_present"] == "yes") else "no"
        exclusion_reason = ""
        confidence = "moderate" if not enroll_rec else "high"
        if include == "no":
            exclusion_reason = "no_grade_12_or_no_acgr"
        if (operating_year, r["school_name"]) in overrides:
            include, exclusion_reason, confidence = overrides[(operating_year, r["school_name"])]
        if "not operating" in r["notes"].lower():
            include = "no"
            exclusion_reason = "school_not_operating"
        rows.append(
            {
                "school_name": r["school_name"],
                "school_code": r["entity_id"],
                "authorizer": r["charter_authorizer"],
                "model_type": r["model_type"],
                "operating_school_year": operating_year,
                "accountability_reporting_year": reporting_year,
                "legal_charter": "yes",
                "grade_9_count": enroll_rec.get("grade_9_count", ""),
                "grade_10_count": enroll_rec.get("grade_10_count", ""),
                "grade_11_count": enroll_rec.get("grade_11_count", ""),
                "grade_12_count": enroll_rec.get("grade_12_count", ""),
                "has_grade_12": has_grade12,
                "acgr_row_exists": "yes" if r["acgr_row_present"] == "yes" else "no",
                "include_in_charter_grad_universe": include,
                "exclusion_reason": exclusion_reason,
                "confidence": confidence,
                "source_file": enroll_rec.get("source_file", Path(r["rate_source_url"]).name or "nevada_charter_school_year_panel_current.csv"),
                "source_url": enroll_rec.get("source_url", r["rate_source_url"]),
                "date_accessed": DATE,
            }
        )

    explicit_non_charters = {
        "Northeastern Nevada Virtual Academy": ("district-operated virtual program", "Elko County School District / district-operated virtual"),
        "NV Learning Academy JR/SR High School": ("district-operated option school", "Clark County School District"),
        "North Star Online School": ("district-operated option school", "Washoe County School District"),
        "NORTH STAR ONLINE SCHOOL": ("district-operated option school", "Washoe County School District"),
        "Davidson Academy": ("non-charter statewide entity", "University Schools"),
        "Independence High School": ("correctional LEA school", "Correctional LEA"),
    }
    seen = {(r["operating_school_year"], r["school_code"], r["school_name"]) for r in rows}
    for r in acgr_export:
        school = r["Group"]
        if school not in explicit_non_charters:
            continue
        grad_year = r["Graduating Class of"].replace("-", "-")[0:4] + "-" + r["Graduating Class of"][7:9]
        if len(grad_year) != 7:
            parts = r["Graduating Class of"].split("-")
            grad_year = f"{parts[0]}-{parts[1][-2:]}"
        key = (grad_year, str(r["Organization Code"]), school)
        if key in seen:
            continue
        rows.append(
            {
                "school_name": school,
                "school_code": str(r["Organization Code"]),
                "authorizer": explicit_non_charters[school][1],
                "model_type": "non_charter_excluded",
                "operating_school_year": grad_year,
                "accountability_reporting_year": short_next_year(grad_year),
                "legal_charter": "no",
                "grade_9_count": "",
                "grade_10_count": "",
                "grade_11_count": "",
                "grade_12_count": "",
                "has_grade_12": "yes" if r["Total Students"] not in {"", "-", None} else "no",
                "acgr_row_exists": "yes",
                "include_in_charter_grad_universe": "no",
                "exclusion_reason": explicit_non_charters[school][0],
                "confidence": "high",
                "source_file": "Nevada_ReportCard_ACGR_school_level_export_2013_14_to_2024_25.csv",
                "source_url": "https://nevadareportcard.nv.gov/DI/main/cohort4yr",
                "date_accessed": DATE,
            }
        )
    rows.sort(key=lambda x: (x["operating_school_year"], x["authorizer"], x["school_name"]))
    return rows


def build_weighted_series(panel: list[dict], summary: list[dict], audit: list[dict], official: list[dict]) -> list[dict]:
    panel_by_year_name = {(r["class_year"], r["school_name"]): r for r in panel}
    by_series_year: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for row in audit:
        by_series_year[(row["class_year"], row["model_type"])].append(row)
    # map audit rows to a release series id
    def audit_series_id(r: dict) -> str | None:
        auth = r["charter_authorizer"]
        model = r["model_type"]
        if auth == "State Public Charter School Authority":
            if model == "statewide_online":
                return "spcsa_virtual"
            if model == "alternative":
                return "spcsa_alternative"
            return "spcsa_brick_mortar"
        if auth == "Clark County School District":
            return "clark_charter_total"
        if auth == "Washoe County School District":
            return "washoe_charter_total"
        return None

    series_line_items: dict[tuple[str, str], list[dict]] = defaultdict(list)
    all_spcsa: dict[str, list[dict]] = defaultdict(list)
    for row in audit:
        sid = audit_series_id(row)
        if sid:
            series_line_items[(row["class_year"], sid)].append(row)
        if row["charter_authorizer"] == "State Public Charter School Authority":
            all_spcsa[row["class_year"]].append(row)

    weighted_lookup = {(r["class_year"], r["series"]): r for r in summary}
    output = []
    required = [
        ("statewide_official", "Statewide Official"),
        ("spcsa_official", "SPCSA Official"),
        ("clark_official", "Clark Official"),
        ("washoe_official", "Washoe Official"),
        ("spcsa_adjusted_charter_only", "SPCSA Charter Only"),
        ("spcsa_brick_mortar", "SPCSA Brick & Mortar"),
        ("spcsa_virtual", "SPCSA Virtual"),
        ("spcsa_alternative", "SPCSA Alternative"),
        ("clark_charter_total", "Clark Charter Total"),
        ("washoe_charter_total", "Washoe Charter Total"),
    ]
    official_map = {
        "State": "statewide_official",
        "State Charters": "spcsa_official",
        "Clark": "clark_official",
        "Washoe": "washoe_official",
    }
    official_lookup = {(r["class_year"], official_map[r["official_row"]]): r for r in official if r["official_row"] in official_map}

    for year in sorted({r["class_year"] for r in summary}):
        for sid, label in required:
            if sid.endswith("_official"):
                off = official_lookup.get((year, sid))
                if not off:
                    continue
                output.append(
                    {
                        "year": year,
                        "series_id": sid,
                        "series_label": label,
                        "cohort": int(off["cohort"]),
                        "graduates": int(off["graduates"]),
                        "weighted_acgr": round(float(off["reported_rate"]), 2),
                        "low_acgr": round(float(off["reported_rate"]), 2),
                        "likely_acgr": round(float(off["reported_rate"]), 2),
                        "high_acgr": round(float(off["reported_rate"]), 2),
                        "included_school_count": "",
                        "suppressed_school_count": "",
                        "current_unweighted_panel_avg": "",
                        "notes": f"Official public rate from Nevada reporting context file; calculated rate from counts was {off['calc_rate']}%",
                    }
                )
                continue
            source_sid = "spcsa_general_ed_virtual" if sid == "spcsa_virtual" else sid
            summ = weighted_lookup.get((year, source_sid))
            if not summ:
                continue
            if sid == "spcsa_adjusted_charter_only":
                items = all_spcsa[year]
            else:
                items = series_line_items.get((year, sid), [])
            low_num = likely_num = high_num = 0.0
            total_cohort = 0.0
            suppressed = 0
            included_names = set()
            for item in items:
                cohort = float(item["cohort"] or 0)
                if cohort <= 0:
                    continue
                panel_row = panel_by_year_name.get((item["class_year"], item["school_name"])) or panel_by_year_name.get((item["class_year"], item["matched_name"]))
                if not panel_row:
                    continue
                included_names.add(item["matched_name"] or item["school_name"])
                total_cohort += cohort
                low = float(panel_row["rate_low"] or panel_row["rate_mid"] or 0)
                mid = float(panel_row["rate_mid"] or panel_row["rate_low"] or 0)
                high = float(panel_row["rate_high"] or panel_row["rate_mid"] or 0)
                low_num += cohort * low
                likely_num += cohort * mid
                high_num += cohort * high
                if panel_row["published_rate_status"] != "numeric":
                    suppressed += 1
            cohort = int(float(summ["cohort_count_used"]))
            graduates = int(float(summ["graduates_used"]))
            output.append(
                {
                    "year": year,
                    "series_id": sid,
                    "series_label": label,
                    "cohort": cohort,
                    "graduates": graduates,
                    "weighted_acgr": round(float(summ["weighted_rate_from_counts"]), 2),
                    "low_acgr": round(low_num / total_cohort, 2) if total_cohort else round(float(summ["weighted_rate_from_counts"]), 2),
                    "likely_acgr": round(likely_num / total_cohort, 2) if total_cohort else round(float(summ["weighted_rate_from_counts"]), 2),
                    "high_acgr": round(high_num / total_cohort, 2) if total_cohort else round(float(summ["weighted_rate_from_counts"]), 2),
                    "included_school_count": int(float(summ["schools_with_counts"])),
                    "suppressed_school_count": suppressed,
                    "current_unweighted_panel_avg": float(summ["current_unweighted_panel_avg"]),
                    "notes": f"Weighted ACGR from school-level cohort and graduate counts. Unweighted comparison was {summ['current_unweighted_panel_avg']}%.",
                }
            )
    return output


def build_lifecycle(universe: list[dict]) -> list[dict]:
    grouped: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for row in universe:
        if row["legal_charter"] != "yes":
            continue
        grouped[(row["school_code"], row["school_name"])].append(row)
    out = []
    for (_, school), rows in sorted(grouped.items(), key=lambda x: x[0][1]):
        authorizer = rows[0]["authorizer"]
        model = rows[0]["model_type"]
        grade12_years = [r["operating_school_year"] for r in rows if r["has_grade_12"] == "yes"]
        acgr_years = [r["operating_school_year"] for r in rows if r["acgr_row_exists"] == "yes"]
        excluded_after = [r for r in rows if r["include_in_charter_grad_universe"] == "no" and r["exclusion_reason"] in {"school_not_operating", "no_grade_12_or_no_acgr"}]
        closure_year = max(acgr_years) if excluded_after and acgr_years else ""
        transition_notes = "; ".join(sorted({r["exclusion_reason"] for r in excluded_after if r["exclusion_reason"]}))[:250]
        out.append(
            {
                "school_name": school,
                "school_code": rows[0]["school_code"],
                "authorizer": authorizer,
                "model_type": model,
                "first_grade12_year": min(grade12_years) if grade12_years else "",
                "last_grade12_year": max(grade12_years) if grade12_years else "",
                "first_acgr_year": min(acgr_years) if acgr_years else "",
                "last_acgr_year": max(acgr_years) if acgr_years else "",
                "closure_year": closure_year,
                "transition_notes": transition_notes,
            }
        )
    return out


def build_validation_exceptions(panel: list[dict]) -> list[dict]:
    out = []
    for r in panel:
        cohort = r["cohort_count"].strip()
        grads = r["graduate_count"].strip()
        rate = r["published_overall_acgr"].strip()
        if not cohort or not grads or not rate:
            continue
        try:
            cohort_n = float(cohort)
            grads_n = float(grads)
            rate_n = float(rate)
        except ValueError:
            continue
        if cohort_n == 0:
            continue
        derived = grads_n / cohort_n * 100
        diff = rate_n - derived
        if abs(diff) > 0.5:
            out.append(
                {
                    "school_name": r["school_name"],
                    "school_code": r["entity_id"],
                    "year": r["class_year"],
                    "published_acgr": round(rate_n, 2),
                    "derived_acgr": round(derived, 2),
                    "difference": round(diff, 2),
                    "cohort": int(cohort_n),
                    "graduates": int(grads_n),
                    "notes": "Published Nevada value retained as official point value; discrepancy flagged for transparency.",
                }
            )
    return out


def build_uncertainty(universe: list[dict], validation: list[dict], panel: list[dict]) -> list[dict]:
    out = []
    panel_lookup = {(r["class_year"], r["school_name"]): r for r in panel}
    for row in universe:
        if row["legal_charter"] == "no":
            out.append(
                {
                    "school_name": row["school_name"],
                    "school_code": row["school_code"],
                    "year": row["operating_school_year"],
                    "uncertainty_type": "legal_status",
                    "low_value": "",
                    "likely_value": "",
                    "high_value": "",
                    "impact_level": "high",
                    "notes": row["exclusion_reason"],
                    "source": row["source_file"],
                }
            )
        elif row["include_in_charter_grad_universe"] == "no":
            out.append(
                {
                    "school_name": row["school_name"],
                    "school_code": row["school_code"],
                    "year": row["operating_school_year"],
                    "uncertainty_type": "no_grade_12",
                    "low_value": "",
                    "likely_value": "",
                    "high_value": "",
                    "impact_level": "medium",
                    "notes": row["exclusion_reason"],
                    "source": row["source_file"],
                }
            )
        panel_row = panel_lookup.get((row["operating_school_year"], row["school_name"]))
        if panel_row and panel_row["published_rate_status"] != "numeric" and panel_row["include_in_sector_total"] == "yes":
            out.append(
                {
                    "school_name": row["school_name"],
                    "school_code": row["school_code"],
                    "year": row["operating_school_year"],
                    "uncertainty_type": "suppression",
                    "low_value": panel_row["rate_low"],
                    "likely_value": panel_row["rate_mid"],
                    "high_value": panel_row["rate_high"],
                    "impact_level": "medium",
                    "notes": panel_row["published_rate_status"],
                    "source": "nevada_charter_school_year_panel_current.csv",
                }
            )
        if panel_row and ("rainshadow" in panel_row["notes"].lower() or "crosswalked" in panel_row["notes"].lower()):
            out.append(
                {
                    "school_name": row["school_name"],
                    "school_code": row["school_code"],
                    "year": row["operating_school_year"],
                    "uncertainty_type": "entity_code_transition",
                    "low_value": "",
                    "likely_value": "",
                    "high_value": "",
                    "impact_level": "medium",
                    "notes": panel_row["notes"][:220],
                    "source": "nevada_charter_school_year_panel_current.csv",
                }
            )
        if panel_row and "ICDA" in panel_row["school_name"] and row["operating_school_year"] >= "2020-21":
            out.append(
                {
                    "school_name": row["school_name"],
                    "school_code": row["school_code"],
                    "year": row["operating_school_year"],
                    "uncertainty_type": "closure_attribution",
                    "low_value": "",
                    "likely_value": "",
                    "high_value": "",
                    "impact_level": "high",
                    "notes": "ICDA closure case; post-closure student outcomes are not attributed back to the charter.",
                    "source": "known_decisions_register.csv",
                }
            )
    for row in validation:
        out.append(
            {
                "school_name": row["school_name"],
                "school_code": row["school_code"],
                "year": row["year"],
                "uncertainty_type": "internal_state_data_conflict",
                "low_value": row["derived_acgr"],
                "likely_value": row["published_acgr"],
                "high_value": row["published_acgr"],
                "impact_level": "medium",
                "notes": row["notes"],
                "source": "data_validation_exceptions.csv",
            }
        )
    return out


def build_release_source_manifest() -> list[dict]:
    rows = load_source_manifest()
    extras = [
        ROOT / "nevada_charter_school_year_panel.csv",
        ROOT / "nevada_charter_panel_decisions.csv",
        ROOT / "nevada_charter_publishable_trend_table.csv",
    ]
    for path in extras:
        h = hashlib.sha256(path.read_bytes()).hexdigest()
        rows.append(
            {
                "relative_path": path.name,
                "source_type": "current_repo_input",
                "file_name": path.name,
                "source_url": "https://github.com/pgavincds/NV_Charter_Grad_rates",
                "date_accessed": DATE,
                "sha256": h,
                "notes": "Current repo input referenced during weighted phase 1 rebuild.",
            }
        )
    return [
        {
            "source_type": r["source_type"],
            "file_name": r["file_name"],
            "source_url": r["source_url"],
            "date_accessed": r["date_accessed"],
            "sha256": r["sha256"],
            "notes": r["notes"],
        }
        for r in rows
    ]


def write_markdown_outputs(weighted: list[dict], lifecycle: list[dict], validation: list[dict], uncertainty: list[dict]) -> None:
    method = RELEASE / "01_Methodology" / "README_weighted_phase1.md"
    findings = RELEASE / "01_Methodology" / "KEY_FINDINGS.md"
    changelog = RELEASE / "07_Change_Log" / "CHANGELOG.md"
    executive = RELEASE / "01_Methodology" / "EXECUTIVE_SUMMARY.md"

    weights = [r for r in weighted if r["series_id"] == "spcsa_adjusted_charter_only"]
    by_year = {r["year"]: r for r in weights}
    peak_grads = max(weights, key=lambda x: x["graduates"])
    intro = f"""# Weighted Phase 1 Method Note

This Phase 1 rebuild uses weighted ACGR as the primary sector measure:

`weighted_acgr = sum(graduates) / sum(cohort)`

The purpose is to describe what happened to the average student in Nevada's charter high-school sector rather than the average school.

## Core rules

1. Legal charter status is separate from school-choice status. District-operated choice, virtual, alternative, and contract schools are excluded even when they appear in the same public reporting universe.
2. Validation Day enrollment files are used to verify grade-12 presence, not as ACGR denominators.
3. Published Nevada ACGR values are preserved as the official point values when they conflict with numerator/denominator-derived rates; those conflicts are flagged in `data_validation_exceptions.csv`.
4. Closed-school cases such as ICDA are documented separately so post-closure student outcomes are not attributed back to the charter without clear public support.
5. Early years were compositionally different: large online and alternative operators dominated the charter high-school denominator before many later brick-and-mortar operators expanded into grade 12.
"""
    method.write_text(intro, encoding="utf-8")

    key = f"""# Key Findings

1. Nevada's charter sector produced more graduates over time as high-school capacity expanded. In this Phase 1 rebuild, the peak charter-only SPCSA graduate count in the weighted file is {peak_grads['graduates']} in {peak_grads['year']}.
2. The high-school sector grew materially, which means early charter outcomes should be read as a composition story as much as a performance story.
3. Virtual schools had an outsized influence in the early years because they carried very large cohorts relative to much of the brick-and-mortar sector.
4. Weighted and unweighted results differ meaningfully. The unweighted school average often overstated sector performance when many small high-performing schools sat alongside a few very large lower-performing schools.
5. The biggest graduate producers are visible in the line-item audit and top-contributor visualizations. The policy story is not only about rates; it is also about who actually graduated students at scale.
6. The assumptions that matter most are legal charter status, grade-12 presence, closure attribution, and whether public numerator/denominator conflicts are silently ignored or surfaced.
"""
    findings.write_text(key, encoding="utf-8")

    change = f"""# Change Log

## Major methodological changes

- Replaced headline unweighted campus averages with weighted ACGR based on graduates and cohort counts.
- Separated operating year from accountability reporting year.
- Added explicit graduation-universe decisions and uncertainty logging.

## Data corrections

- Carried forward explicit exclusions for non-charter rows such as Northeastern Nevada Virtual Academy, NV Learning Academy, North Star Online, Davidson, and Independence.
- Applied release-time overrides for YWLA and pre-grade-12 Explore Academy years.

## Inclusion and exclusion changes

- Documented non-charter exclusions directly in the universe table.
- Preserved known charter cases such as Quest and ICDA in line with vetted findings.

## Closure cases

- ICDA remains included for reported cohorts only. No synthetic post-closure outcomes were created.

## Remaining open questions

- Some early years still lack direct Validation Day grade counts.
- White Pine district-operated virtual and contract cases are acknowledged but not fully reconstructed as separate line items in this Phase 1 release.
- Public numerator/denominator conflicts remain visible in the validation exceptions file.

## Future research

- Add earlier grade-by-grade enrollment evidence for pre-2021 years.
- Improve school-lifecycle closure dating for legacy charters.
- Extend the public website to let users switch between weighted and unweighted series interactively.
"""
    changelog.write_text(change, encoding="utf-8")

    exec_text = f"""# Executive Summary

This Phase 1 release rebuilds Nevada charter graduation-rate reporting around student-weighted ACGR rather than average-of-school rates.

- Earliest weighted SPCSA charter-only ACGR in this release: {weights[0]['weighted_acgr']}% in {weights[0]['year']}
- Latest weighted SPCSA charter-only ACGR in this release: {weights[-1]['weighted_acgr']}% in {weights[-1]['year']}
- Data validation exceptions flagged: {len(validation)}
- Uncertainty register entries created: {len(uncertainty)}
- Lifecycle rows created: {len(lifecycle)}

The release is designed for policymakers, journalists, and researchers who need a transparent, reproducible starting point rather than a black-box narrative.
"""
    executive.write_text(exec_text, encoding="utf-8")


def write_public_pages() -> None:
    web_dir = RELEASE / "06_Public_Website"
    web_dir.mkdir(parents=True, exist_ok=True)
    overview = """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Nevada Charter Graduation Rates Weighted Rebuild</title>
  <style>
    body { font-family: Georgia, 'Times New Roman', serif; background: #f5f1e8; color: #1f2a2e; margin: 0; padding: 2rem; }
    main { max-width: 960px; margin: 0 auto; }
    a { color: #0f5c63; }
  </style>
</head>
<body>
  <main>
    <h1>Nevada Charter Graduation Rates Weighted Rebuild</h1>
    <p>This Phase 1 release rebuilds the Nevada charter graduation-rate series using weighted ACGR, documented graduation-universe rules, and explicit source manifests.</p>
    <ul>
      <li><a href="../03_Clean_Data/weighted_sector_series.csv">Weighted sector series</a></li>
      <li><a href="../03_Clean_Data/graduation_universe.csv">Graduation universe</a></li>
      <li><a href="../05_Visualizations/weighted_acgr_trend.html">Weighted ACGR trend</a></li>
      <li><a href="../05_Visualizations/graduates_produced_by_year.html">Graduates produced by year</a></li>
      <li><a href="../01_Methodology/KEY_FINDINGS.md">Key findings</a></li>
    </ul>
  </main>
</body>
</html>
"""
    (web_dir / "index.html").write_text(overview, encoding="utf-8")


def copy_source_data() -> None:
    src_dir = RELEASE / "02_Source_Data"
    src_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(HANDOFF / "01_SOURCE_DATA" / "source_manifest.csv", src_dir / "source_manifest_original.csv")
    shutil.copy2(HANDOFF / "01_SOURCE_DATA" / "ACGR_exports" / "Nevada_ReportCard_ACGR_school_level_export_2013_14_to_2024_25.csv", src_dir)
    for path in (HANDOFF / "01_SOURCE_DATA" / "Validation_Day_enrollment").glob("*.xlsx"):
        shutil.copy2(path, src_dir / path.name)


def write_release():
    panel = load_current_panel()
    summary = load_weighted_summary()
    audit = load_weighted_audit()
    official = load_official_context()
    acgr_export = read_csv_dict(HANDOFF / "01_SOURCE_DATA" / "ACGR_exports" / "Nevada_ReportCard_ACGR_school_level_export_2013_14_to_2024_25.csv", skip_report_preamble=True)
    enroll = load_enrollment_counts()

    if RELEASE.exists():
        shutil.rmtree(RELEASE)
    for sub in ["01_Methodology", "02_Source_Data", "03_Clean_Data", "04_Audit_Files", "05_Visualizations", "06_Public_Website", "07_Change_Log"]:
        (RELEASE / sub).mkdir(parents=True, exist_ok=True)

    universe = build_universe(panel, acgr_export, enroll)
    weighted = build_weighted_series(panel, summary, audit, official)
    lifecycle = build_lifecycle(universe)
    validation = build_validation_exceptions(panel)
    uncertainty = build_uncertainty(universe, validation, panel)
    source_manifest = build_release_source_manifest()

    weighted_fields = [
        "year", "series_id", "series_label", "cohort", "graduates", "weighted_acgr",
        "low_acgr", "likely_acgr", "high_acgr", "included_school_count", "suppressed_school_count",
        "current_unweighted_panel_avg", "notes",
    ]
    write_csv(RELEASE / "03_Clean_Data" / "weighted_sector_series.csv", weighted_fields, weighted)
    write_csv(RELEASE / "03_Clean_Data" / "graduation_universe.csv", list(universe[0].keys()), universe)
    write_csv(RELEASE / "03_Clean_Data" / "school_lifecycle_report.csv", list(lifecycle[0].keys()), lifecycle)
    write_csv(RELEASE / "04_Audit_Files" / "data_validation_exceptions.csv", list(validation[0].keys()) if validation else ["school_name","school_code","year","published_acgr","derived_acgr","difference","cohort","graduates","notes"], validation)
    write_csv(RELEASE / "04_Audit_Files" / "uncertainty_register.csv", list(uncertainty[0].keys()), uncertainty)
    write_csv(RELEASE / "02_Source_Data" / "source_manifest.csv", list(source_manifest[0].keys()), source_manifest)
    write_csv(RELEASE / "04_Audit_Files" / "weighted_line_items_audit_release.csv", list(audit[0].keys()), audit)

    copy_source_data()
    write_markdown_outputs(weighted, lifecycle, validation, uncertainty)
    write_public_pages()
    build_visualizations(RELEASE)

    zip_path = ROOT / "release" / "NV_Charter_Grad_Rates_Weighted_Rebuild_v1.zip"
    if zip_path.exists():
        zip_path.unlink()
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for path in RELEASE.rglob("*"):
            if path.is_file():
                zf.write(path, path.relative_to(RELEASE.parent))

    manifest = {
        "release_dir": str(RELEASE),
        "zip_path": str(zip_path),
        "files": sorted(str(p.relative_to(RELEASE)) for p in RELEASE.rglob("*") if p.is_file()),
    }
    (RELEASE / "07_Change_Log" / "release_manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")


if __name__ == "__main__":
    write_release()
